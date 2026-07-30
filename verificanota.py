import os
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime

# ========================= CONFIGURAÇÕES =========================
# Planilha de entrada original — a que VOCÊ atualiza (mesma do main.py).
ARQUIVO_LISTA = r"C:\Users\5829097\OneDrive\Onedrive - GPA\Área de Trabalho\pegarnotas.xlsx"
PASTA_DESTINO = r"C:\Users\5829097\OneDrive\Onedrive - GPA\Área de Trabalho\NOTAS RECLASSIFICAÇÃO"

# Mesmas colunas/linha inicial do main.py
ABA = None
COL_OCOR = "H"
COL_NOTA = "C"
LINHA_INICIO = 2

# ---- SAÍDAS ----
# Nova base SÓ com as notas faltantes, no MESMO formato que o main.py lê
# (coluna C já com o sufixo " (2)" quando duplicada; coluna H com a
# ocorrência). A PEGARNOTAS original NÃO é alterada. Ao final, o terminal
# mostra a linha pronta para colar no ARQUIVO_LISTA do main.py.
ARQUIVO_FALTANTES = r"C:\Users\5829097\Downloads\PEGARNOTAS_FALTANTES.xlsx"

# Planilha de verificação detalhada (todas as linhas, com situação)
ARQUIVO_VERIFICACAO = os.path.join(PASTA_DESTINO, "_verificacao_nfse.xlsx")
# =================================================================


def limpar(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def extrair_ocorrencia(valor):
    """Mesma regra do main.py: '4800516664 - FIN0493074' → '4800516664'."""
    valor = limpar(valor)
    m = re.match(r"^(\d+)\s*-\s*FIN\d*", valor, re.IGNORECASE)
    if m:
        return m.group(1)
    return valor


def pdf_valido(caminho):
    """Verifica se o arquivo realmente começa com a assinatura de um PDF (%PDF-)."""
    try:
        if not os.path.exists(caminho) or os.path.getsize(caminho) == 0:
            return False
        with open(caminho, "rb") as f:
            assinatura = f.read(5)
        return assinatura == b"%PDF-"
    except Exception:
        return False


def ler_lista_com_nomes_unicos():
    """
    Lê a PEGARNOTAS e atribui a cada linha um NOME DE ARQUIVO ÚNICO:
      1ª vez que a nota aparece → '12345'
      2ª vez                    → '12345 (2)'
      3ª vez                    → '12345 (3)'  ... e assim por diante.
    Assim cada linha da planilha corresponde a exatamente UM arquivo esperado
    na pasta, mesmo quando o número da nota se repete (notas diferentes com o
    mesmo número).
    """
    if not os.path.exists(ARQUIVO_LISTA):
        print(f"ERRO: planilha de entrada não encontrada em:\n  {ARQUIVO_LISTA}")
        return []

    wb = openpyxl.load_workbook(ARQUIVO_LISTA, data_only=True)
    ws = wb[ABA] if ABA else wb.active

    registros = []
    contagem_nota = defaultdict(int)

    for i in range(LINHA_INICIO, ws.max_row + 1):
        ocorrencia_bruta = limpar(ws[f"{COL_OCOR}{i}"].value)
        ocorrencia       = extrair_ocorrencia(ocorrencia_bruta)
        num_nota         = limpar(ws[f"{COL_NOTA}{i}"].value)

        # mesma regra do main.py: linha sem ocorrência é pulada
        if not ocorrencia:
            continue

        contagem_nota[num_nota] += 1
        vez = contagem_nota[num_nota]
        nome_unico = num_nota if vez == 1 else f"{num_nota} ({vez})"

        registros.append({
            "linha": i,
            "ocorrencia": ocorrencia,
            "ocorrencia_bruta": ocorrencia_bruta,
            "nota": num_nota,
            "nome_unico": nome_unico,     # nome do arquivo esperado (sem .pdf)
            "duplicada": vez > 1,
        })

    duplicadas = sum(1 for r in registros if r["duplicada"])
    print(f"{len(registros)} linha(s) lidas de PEGARNOTAS.xlsx "
          f"({duplicadas} com número de nota repetido).")
    return registros


def verificar(registros):
    """
    Confere a pasta pelo NOME ÚNICO de cada linha:
      - OK             → '{nome_unico}.pdf' está na pasta e é válido
      - PDF CORROMPIDO → arquivo existe mas não tem assinatura de PDF
      - FALTANDO       → arquivo '{nome_unico}.pdf' não está na pasta
    Observação: uma linha duplicada NUNCA "empresta" o PDF da primeira —
    ela procura pelo próprio nome com sufixo (ex.: '12345 (2).pdf').
    """
    resultados = []
    for r in registros:
        caminho = os.path.join(PASTA_DESTINO, f"{r['nome_unico']}.pdf")
        existe  = os.path.exists(caminho)
        valido  = pdf_valido(caminho) if existe else False

        if valido:
            situacao = "OK"
            detalhe  = f"{r['nome_unico']}.pdf presente e válido"
        elif existe:
            situacao = "PDF CORROMPIDO"
            detalhe  = f"{r['nome_unico']}.pdf existe mas não é um PDF válido"
        else:
            situacao = "FALTANDO"
            if r["duplicada"]:
                detalhe = (f"Esperado '{r['nome_unico']}.pdf' — nota duplicada; "
                           "o download anterior provavelmente sobrescreveu com o mesmo nome")
            else:
                detalhe = f"Esperado '{r['nome_unico']}.pdf' — não está na pasta"

        resultados.append({**r, "situacao": situacao, "detalhe": detalhe})
    return resultados


def listar_pdfs_orfaos(registros):
    """PDFs na pasta que não correspondem a nenhum nome único esperado."""
    esperados = {r["nome_unico"] for r in registros}
    orfaos = []
    try:
        for nome in os.listdir(PASTA_DESTINO):
            base, ext = os.path.splitext(nome)
            if ext.lower() == ".pdf" and base not in esperados:
                orfaos.append(nome)
    except Exception:
        pass
    return sorted(orfaos)


def gerar_base_faltantes(faltantes):
    """
    Gera a nova base para reprocessamento em ARQUIVO SEPARADO (a PEGARNOTAS
    original não é tocada), no mesmo layout que o main.py lê:
      - Coluna C: o NOME ÚNICO (nota já com o sufixo ' (2)' quando duplicada).
        Como o main.py salva o PDF como '{coluna C}.pdf', o arquivo já sai com
        o nome certo — sem alterar o main.py nem o nix_material.py.
      - Coluna H: a ocorrência (formato bruto original, que o main.py já sabe
        tratar com extrair_ocorrencia).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FALTANTES"

    # Cabeçalhos nas mesmas colunas usadas pelos scripts
    ws[f"{COL_NOTA}1"] = "NOTA (nome do arquivo)"
    ws[f"{COL_OCOR}1"] = "OCORRÊNCIA"
    ws[f"{COL_NOTA}1"].font = Font(bold=True)
    ws[f"{COL_OCOR}1"].font = Font(bold=True)

    for idx, r in enumerate(faltantes, LINHA_INICIO):
        ws[f"{COL_NOTA}{idx}"] = r["nome_unico"]
        # usa a ocorrência bruta original se existir (o main.py extrai sozinho)
        ws[f"{COL_OCOR}{idx}"] = r["ocorrencia_bruta"] or r["ocorrencia"]

    ws.column_dimensions[COL_NOTA].width = 25
    ws.column_dimensions[COL_OCOR].width = 28

    try:
        wb.save(ARQUIVO_FALTANTES)
    except PermissionError:
        print("\n⚠️  FECHE o PEGARNOTAS_FALTANTES.xlsx no Excel e rode de novo — "
              "não foi possível sobrescrevê-lo aberto.")
        return None

    return ARQUIVO_FALTANTES


def gerar_planilha_verificacao(resultados, orfaos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verificação"

    fill_verde    = PatternFill("solid", fgColor="C6EFCE")
    fill_amarelo  = PatternFill("solid", fgColor="FFEB9C")
    fill_vermelho = PatternFill("solid", fgColor="FFC7CE")
    fill_titulo   = PatternFill("solid", fgColor="1F3864")
    fill_cabec    = PatternFill("solid", fgColor="2E75B6")

    fonte_titulo = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    fonte_cabec  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    fonte_normal = Font(name="Calibri", size=11)

    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    centro = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Verificação NFS-e (base: PEGARNOTAS.xlsx) — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = fonte_titulo
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = centro
    ws.row_dimensions[1].height = 30

    cabecalhos = ["Linha", "Ocorrência", "Nota Fiscal", "Arquivo esperado",
                  "Duplicada?", "Situação", "Detalhe"]
    for col, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=2, column=col, value=cab)
        cell.font = fonte_cabec
        cell.fill = fill_cabec
        cell.alignment = centro
        cell.border = borda
    ws.row_dimensions[2].height = 22

    for row_idx, r in enumerate(resultados, 3):
        if r["situacao"] == "OK":
            fill = fill_verde
        elif r["situacao"] == "FALTANDO":
            fill = fill_amarelo
        else:
            fill = fill_vermelho

        valores = [r["linha"], r["ocorrencia"], r["nota"],
                   f"{r['nome_unico']}.pdf",
                   "SIM" if r["duplicada"] else "",
                   r["situacao"], r["detalhe"]]
        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = fonte_normal
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="left" if col_idx == 7 else "center",
                vertical="center"
            )
            cell.border = borda

    for col, larg in enumerate([8, 20, 18, 24, 12, 16, 60], 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    if orfaos:
        ws2 = wb.create_sheet("PDFs fora da lista")
        ws2["A1"] = "Arquivos .pdf presentes na pasta mas sem linha correspondente"
        ws2["A1"].font = Font(bold=True)
        for idx, nome in enumerate(orfaos, 3):
            ws2.cell(row=idx, column=1, value=nome)
        ws2.column_dimensions["A"].width = 40

    wb.save(ARQUIVO_VERIFICACAO)
    return ARQUIVO_VERIFICACAO


def imprimir_resumo(resultados, orfaos, caminho_faltantes):
    ok          = [r for r in resultados if r["situacao"] == "OK"]
    faltando    = [r for r in resultados if r["situacao"] == "FALTANDO"]
    corrompidos = [r for r in resultados if r["situacao"] == "PDF CORROMPIDO"]
    duplicadas  = [r for r in resultados if r["duplicada"]]

    print("\n" + "=" * 64)
    print("       VERIFICAÇÃO — PEGARNOTAS.xlsx × PASTA DE NOTAS")
    print("=" * 64)
    print(f"  Linhas na planilha              : {len(resultados)}")
    print(f"  Linhas com nota duplicada       : {len(duplicadas)}")
    print("-" * 64)
    print(f"  ✔  OK (arquivo próprio na pasta): {len(ok)}")
    print(f"  ⚠  FALTANDO                     : {len(faltando)}")
    print(f"  ✘  PDF corrompido               : {len(corrompidos)}")
    print(f"  PDFs na pasta fora da lista     : {len(orfaos)}")
    print("=" * 64)

    if not faltando and not corrompidos:
        print("\n  ✅ TUDO CERTO: cada linha da planilha tem seu próprio PDF na pasta.")
    else:
        print(f"\n  ❌ {len(faltando) + len(corrompidos)} linha(s) precisam ser (re)baixadas:")

    if faltando:
        print("\n  FALTANDO:")
        for r in faltando:
            dup = " [DUPLICADA]" if r["duplicada"] else ""
            print(f"    Linha {str(r['linha']):>4} | Ocorrência {r['ocorrencia']} | "
                  f"Arquivo esperado: {r['nome_unico']}.pdf{dup}")

    if corrompidos:
        print("\n  PDF CORROMPIDO:")
        for r in corrompidos:
            print(f"    Linha {str(r['linha']):>4} | Ocorrência {r['ocorrencia']} | "
                  f"{r['nome_unico']}.pdf")

    if orfaos:
        print("\n  PDFs na pasta sem linha na planilha:")
        for nome in orfaos:
            print(f"    {nome}")

    if caminho_faltantes:
        print("\n" + "-" * 64)
        print("  Nova base com as faltantes gerada (PEGARNOTAS original intacta).")
        print("  PRÓXIMOS PASSOS:")
        print("  1. No main.py, substitua a linha do ARQUIVO_LISTA por:")
        print()
        print(f'     ARQUIVO_LISTA = r"{caminho_faltantes}"')
        print()
        print("  2. Rode main.py e depois nix_material.py normalmente.")
        print("     (Os PDFs já sairão com o nome único, ex.: '12345 (2).pdf')")
        print("  3. Ao terminar, volte o ARQUIVO_LISTA para a PEGARNOTAS original")
        print("     e rode esta verificação de novo para confirmar que fechou.")
        print("-" * 64)

    print("=" * 64 + "\n")


def main():
    registros = ler_lista_com_nomes_unicos()
    if not registros:
        return

    resultados = verificar(registros)
    orfaos     = listar_pdfs_orfaos(registros)

    # Linhas que precisam ser (re)baixadas: faltantes + corrompidas
    pendentes = [r for r in resultados if r["situacao"] in ("FALTANDO", "PDF CORROMPIDO")]

    caminho_faltantes = None
    if pendentes:
        caminho_faltantes = gerar_base_faltantes(pendentes)

    imprimir_resumo(resultados, orfaos, caminho_faltantes)
    caminho_verif = gerar_planilha_verificacao(resultados, orfaos)
    print(f"Planilha de verificação salva em: {caminho_verif}")
    if caminho_faltantes:
        print(f"Base de faltantes salva em      : {caminho_faltantes}")


if __name__ == "__main__":
    main()