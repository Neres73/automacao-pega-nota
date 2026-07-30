import os
import re
import time
import shutil
import win32com.client
import win32gui
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pyautogui
import pyperclip
from pywinauto import Desktop
from datetime import datetime, timedelta

# ========================= CONFIGURAÇÕES =========================
ARQUIVO_LISTA = r"C:\Users\5829097\OneDrive\Onedrive - GPA\Área de Trabalho\pegarnotas.xlsx"


PASTA_DESTINO = r"C:\Users\5829097\OneDrive\Onedrive - GPA\Área de Trabalho\NOTAS RECLASSIFICAÇÃO"

ABA = None
COL_OCOR = "H"
COL_NOTA = "C"
LINHA_INICIO = 2

TRES_PONTOS = (800, 160)
SETAS_ATE_SALVAR = 2

# --- Tempos ---
# ESPERA_RENDER: espera o PDF renderizar antes de abrir o menu dos 3 pontos.
# É difícil "pollar" esse render de forma confiável, então segue fixo — mas se
# o seu ambiente for consistente, teste baixar para 2.0.
ESPERA_RENDER = 3.0
# ESPERA_MENU: tempo para o menu dos 3 pontos abrir. 1.5 era generoso; 0.8 costuma bastar.
ESPERA_MENU   = 0.8

# --- Espera ativa (polling) da gravação do PDF ---
# Em vez de dormir um tempo fixo após clicar em "Salvar", verificamos o arquivo
# em intervalos curtos e seguimos assim que ele estiver pronto (válido e com
# tamanho estável). TIMEOUT_PDF é só o teto de segurança.
TIMEOUT_PDF   = 8.0
INTERVALO_POLL = 0.2

# --- Reuso da tela de seleção (ganho estrutural, EXPERIMENTAL) ---
# Se True, a transação /nNIXNFSE só é aberta na 1ª nota; nas demais o script
# tenta voltar (F3) para a tela de seleção e apenas troca a ocorrência, evitando
# reiniciar a transação e reabrir a árvore a cada linha.
# ATENÇÃO: o número de "voltas" (F3) depende de como o NIXNFSE se comporta ao
# retornar da visão do PDF e do grid. TESTE com 2-3 notas antes de rodar tudo.
# Em caso de qualquer dúvida, deixe em False (comportamento idêntico ao original).
REUSAR_TELA_SELECAO = False
VOLTAS_ATE_SELECAO = 2   # quantos F3 para sair da visão do PDF até a tela de seleção

# --- Progresso no terminal ---
# A cada quantas notas imprimir a linha de PARCIAL (resumo acumulado + ETA).
PARCIAL_A_CADA = 10

# Reduz a pausa interna que o pyautogui insere entre CADA chamada (padrão 0.1s).
pyautogui.PAUSE = 0.02

# Textos possíveis do combo de tipo de arquivo na janela "Salvar como".
TIPOS_PDF_ACEITOS = ["PDF Files", "PDF (*.pdf)", "Arquivos PDF", "PDF", "Todos os arquivos (*.*)"]
# =================================================================


def limpar(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def extrair_ocorrencia(valor):
    """
    Algumas linhas trazem a ocorrência num formato composto, ex:
      '4800516664 - FIN0493074'  ou  '4800512991-FIN0486244'
    O SAP não reconhece esse formato no campo de busca e isso causa o erro
    'invalid argument' (SAP Frontend Server) mais adiante no fluxo.
    Aqui extraímos só o número principal antes do '- FIN'.
    """
    valor = limpar(valor)
    m = re.match(r"^(\d+)\s*-\s*FIN\d*", valor, re.IGNORECASE)
    if m:
        return m.group(1)
    return valor


def conectar_sap():
    sap_gui = win32com.client.GetObject("SAPGUI")
    app = sap_gui.GetScriptingEngine
    connection = app.Children(0)
    session = connection.Children(0)
    return session


def pdf_valido(caminho):
    """Verifica se o arquivo realmente começa com a assinatura de um PDF (%PDF-).
    Isso evita 'falsos positivos' de arquivos salvos com extensão .pdf mas que
    na verdade são HTML/MHT/lixo (o erro clássico do Acrobat 'tipo não suportado')."""
    try:
        if not os.path.exists(caminho) or os.path.getsize(caminho) == 0:
            return False
        with open(caminho, "rb") as f:
            assinatura = f.read(5)
        return assinatura == b"%PDF-"
    except Exception:
        return False


def esperar_pdf_pronto(caminho, timeout=TIMEOUT_PDF, intervalo=INTERVALO_POLL):
    """
    Espera ATIVA: retorna assim que o PDF existir, for válido e tiver o tamanho
    estável (igual em duas leituras seguidas) — evita pegar o arquivo pela metade.
    Normalmente termina bem antes do timeout, substituindo o antigo sleep fixo.
    Retorna True se ficou pronto dentro do prazo, False caso contrário.
    """
    fim = time.time() + timeout
    tam_anterior = -1
    estavel = 0
    while time.time() < fim:
        if pdf_valido(caminho):
            try:
                tam = os.path.getsize(caminho)
            except Exception:
                tam = -1
            if tam == tam_anterior and tam > 0:
                estavel += 1
                if estavel >= 2:
                    return True
            else:
                estavel = 0
                tam_anterior = tam
        time.sleep(intervalo)
    return False


def selecionar_tipo_pdf(dlg):
    """Tenta forçar o combo de 'Tipo de arquivo' da janela Salvar como para PDF.
    Se não achar o combo ou nenhuma opção compatível, ignora silenciosamente
    (fallback: o nome do arquivo com .pdf pode já bastar em alguns diálogos)."""
    try:
        combos = dlg.children(class_name="ComboBox")
        if not combos:
            # alguns diálogos usam ComboBoxEx32
            combos = dlg.children(class_name="ComboBoxEx32")
        for combo in combos:
            try:
                itens = combo.item_texts()
            except Exception:
                continue
            for alvo in TIPOS_PDF_ACEITOS:
                for item in itens:
                    if alvo.lower() in item.lower():
                        combo.select(item)
                        time.sleep(0.2)
                        return True
    except Exception as e:
        print(f"  aviso: não foi possível checar/selecionar tipo de arquivo: {e}")
    return False


def salvar_via_pywinauto(caminho_completo):
    """
    Trata a janela 'Salvar como' do Windows em background —
    sem mover o mouse nem precisar de foco do teclado.
    NÃO espera a gravação aqui: quem aguarda o PDF ficar pronto é o
    esperar_pdf_pronto() no fluxo principal (espera ativa).
    """
    try:
        desktop = Desktop(backend="win32")

        # aguarda a janela "Salvar como" aparecer (até 10s)
        dlg = desktop.window(title_re="Salvar como|Save As|Salvar Saída de Impressão")
        dlg.wait("visible", timeout=10)
        time.sleep(0.15)   # pequena folga; o wait acima já garante que apareceu

        # tenta garantir que o tipo de arquivo selecionado seja PDF
        selecionar_tipo_pdf(dlg)

        # encontra o campo "Nome do arquivo" e cola o caminho
        try:
            edit = dlg.child_window(class_name="Edit")
            edit.set_edit_text(caminho_completo)
        except Exception:
            # fallback: último campo Edit da janela
            edits = dlg.children(class_name="Edit")
            if edits:
                edits[-1].set_edit_text(caminho_completo)

        time.sleep(0.2)

        # clica no botão Salvar
        try:
            dlg.child_window(title_re="&Salvar|Salvar|Save",
                             class_name="Button").click()
        except Exception:
            dlg.type_keys("{ENTER}")

        return True

    except Exception as e:
        print(f"  pywinauto erro: {e} — tentando fallback pyautogui...")
        # fallback: se pywinauto falhar, usa pyautogui como antes
        pyperclip.copy(caminho_completo)
        pyautogui.hotkey("ctrl", "a")
        time.sleep(0.2)
        pyautogui.hotkey("ctrl", "v")
        time.sleep(0.3)
        pyautogui.press("enter")
        return False


def salvar_nfse(session, caminho_completo):
    """
    Clica nos 3 pontos (pyautogui — rápido, ~1s),
    navega o menu por teclado e trata o Salvar como em background (pywinauto).
    A espera pela gravação do arquivo é feita fora daqui, via esperar_pdf_pronto().
    """
    # 1) clica nos 3 pontos — único momento que move o mouse
    pyautogui.click(*TRES_PONTOS)
    time.sleep(ESPERA_MENU)

    # 2) navega no menu por teclado
    for _ in range(SETAS_ATE_SALVAR):
        pyautogui.press("down")
        time.sleep(0.2)
    pyautogui.press("enter")
    time.sleep(0.4)   # só espera o menu fechar; pywinauto aguarda a janela

    # 3) trata a janela "Salvar como" em background (não bloqueia esperando o disco)
    salvar_via_pywinauto(caminho_completo)


def abrir_selecao(session):
    """Entra na transação /nNIXNFSE e abre o nó da árvore de seleção."""
    session.findById("wnd[0]/tbar[0]/okcd").Text = "/nNIXNFSE"
    session.findById("wnd[0]").sendVKey(0)
    session.findById(
        "wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell"
    ).doubleClickNode("F00003")


def voltar_para_selecao(session):
    """
    (EXPERIMENTAL) Volta da visão do PDF/grid até a tela de seleção usando F3,
    evitando reiniciar a transação. Ajuste VOLTAS_ATE_SELECAO conforme o
    comportamento do seu NIXNFSE. Retorna True se conseguiu chegar num estado
    onde o campo de ocorrência existe.
    """
    for _ in range(VOLTAS_ATE_SELECAO):
        try:
            session.findById("wnd[0]").sendVKey(3)  # F3 = Voltar
            time.sleep(0.3)
        except Exception:
            break
    # valida se estamos na tela de seleção (campo da ocorrência presente)
    try:
        session.findById(
            "wnd[0]/usr/ssubSUB_SEL:/MIDAS/NIXNFSE_COCKPIT:0001/txtS_OCOR-LOW"
        )
        return True
    except Exception:
        return False


def limpar_extras(pasta, num_nota):
    pasta_files = os.path.join(pasta, f"{num_nota}_files")
    if os.path.isdir(pasta_files):
        shutil.rmtree(pasta_files, ignore_errors=True)
    try:
        for nome in os.listdir(pasta):
            base, ext = os.path.splitext(nome)
            if base == num_nota and ext.lower() != ".pdf":
                try:
                    os.remove(os.path.join(pasta, nome))
                except Exception:
                    pass
    except Exception:
        pass


def formatar_duracao(segundos):
    """Formata segundos como '1h23m', '5m40s' ou '37s'."""
    segundos = int(segundos)
    h, resto = divmod(segundos, 3600)
    m, s = divmod(resto, 60)
    if h:
        return f"{h}h{m:02d}m"
    if m:
        return f"{m}m{s:02d}s"
    return f"{s}s"


def imprimir_parcial(processadas, total, ok, erros, materiais, inicio):
    """Linha de PARCIAL: acumulado + tempo médio + estimativa de término."""
    decorrido = time.time() - inicio
    media = decorrido / processadas if processadas else 0
    restantes = total - processadas
    eta_seg = media * restantes
    hora_fim = (datetime.now() + timedelta(seconds=eta_seg)).strftime("%H:%M")
    pct = (processadas / total * 100) if total else 0

    print(f"  ── PARCIAL {processadas}/{total} ({pct:.0f}%) | "
          f"✔ {ok} OK | ✘ {erros} erro(s) | {materiais} material | "
          f"média {media:.1f}s/nota | restam ~{formatar_duracao(eta_seg)} "
          f"(término ~{hora_fim}) ──")


def main():
    os.makedirs(PASTA_DESTINO, exist_ok=True)

    wb_lista = openpyxl.load_workbook(ARQUIVO_LISTA, data_only=True)
    ws_lista = wb_lista[ABA] if ABA else wb_lista.active

    # Pré-conta as linhas válidas para mostrar o total real no progresso
    ultima_linha = ws_lista.max_row
    linhas_validas = []
    for i in range(LINHA_INICIO, ultima_linha + 1):
        if limpar(ws_lista[f"{COL_OCOR}{i}"].value):
            linhas_validas.append(i)
    total = len(linhas_validas)

    print("=" * 64)
    print(f"  INÍCIO — {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"  Planilha : {ARQUIVO_LISTA}")
    print(f"  Destino  : {PASTA_DESTINO}")
    print(f"  Notas a processar: {total}")
    print("=" * 64)

    session = conectar_sap()
    session.findById("wnd[0]").maximize()

    dados_relatorio = []
    selecao_pronta = False   # controla o reuso da tela de seleção

    inicio_geral = time.time()
    processadas = 0
    cont_ok = cont_erro = cont_material = 0

    for i in linhas_validas:
        ocorrencia_bruta = limpar(ws_lista[f"{COL_OCOR}{i}"].value)
        ocorrencia       = extrair_ocorrencia(ocorrencia_bruta)
        num_nota         = limpar(ws_lista[f"{COL_NOTA}{i}"].value)

        nome_arquivo = f"{num_nota}.pdf"
        caminho      = os.path.join(PASTA_DESTINO, nome_arquivo)

        processadas += 1
        inicio_nota = time.time()
        prefixo = f"[{processadas}/{total}] Linha {i}"
        print(f"{prefixo} | Ocorrência {ocorrencia} | Nota {num_nota} → processando...")

        try:
            # --- Abre / reaproveita a tela de seleção ---
            if REUSAR_TELA_SELECAO and selecao_pronta:
                if not voltar_para_selecao(session):
                    # não conseguiu voltar limpo: reinicia a transação por segurança
                    abrir_selecao(session)
            else:
                abrir_selecao(session)
            selecao_pronta = True

            # --- Preenche filtros e executa ---
            session.findById(
                "wnd[0]/usr/ssubSUB_SEL:/MIDAS/NIXNFSE_COCKPIT:0001/ctxtS_DTINTE-LOW"
            ).Text = ""
            session.findById(
                "wnd[0]/usr/ssubSUB_SEL:/MIDAS/NIXNFSE_COCKPIT:0001/txtS_OCOR-LOW"
            ).Text = ocorrencia
            session.findById("wnd[0]/tbar[1]/btn[8]").press()

            try:
                grid = session.findById("wnd[0]/usr/cntlCONT_COCKPIT/shellcont/shell")
                row_count = grid.RowCount
            except Exception:
                row_count = 0

            if row_count == 0:
                dados_relatorio.append({
                    "linha": i, "ocorrencia": ocorrencia, "nota": num_nota,
                    "tipo": "Material", "status": "Não encontrada no SAP",
                    "obs": "NFS-e não localizada para esta ocorrência"
                })
                cont_material += 1
                print(f"{prefixo} → MATERIAL (não encontrada no SAP) "
                      f"[{time.time() - inicio_nota:.1f}s]")
                # Nota: neste caminho não abrimos o PDF, então o estado da tela é
                # diferente. Se estiver usando REUSAR_TELA_SELECAO, forçamos uma
                # reabertura limpa na próxima nota para não errar as "voltas".
                if REUSAR_TELA_SELECAO:
                    selecao_pronta = False
                if processadas % PARCIAL_A_CADA == 0:
                    imprimir_parcial(processadas, total, cont_ok, cont_erro,
                                     cont_material, inicio_geral)
                continue

            print(f"{prefixo} → encontrada no SAP, abrindo PDF...")

            grid.currentCellColumn = "OCORRENCIA"
            grid.selectedRows = "0"
            grid.clickCurrentCell()

            try:
                session.findById("wnd[1]").sendVKey(0)
            except Exception:
                pass

            session.findById("wnd[0]/tbar[1]/btn[7]").press()
            time.sleep(ESPERA_RENDER)

            print(f"{prefixo} → salvando {nome_arquivo}...")
            salvar_nfse(session, caminho)
            esperar_pdf_pronto(caminho)          # espera ATIVA (substitui o sleep fixo)
            limpar_extras(PASTA_DESTINO, num_nota)

            if pdf_valido(caminho):
                status = "OK"
                obs    = caminho
                if ocorrencia != ocorrencia_bruta:
                    obs += f"  (ocorrência original: {ocorrencia_bruta} → usada: {ocorrencia})"
                cont_ok += 1
                print(f"{prefixo} → ✔ {nome_arquivo} salvo "
                      f"[{time.time() - inicio_nota:.1f}s]")
            else:
                # arquivo não é um PDF de verdade (ou não existe) — remove pra não confundir depois
                if os.path.exists(caminho):
                    try:
                        os.remove(caminho)
                    except Exception:
                        pass
                status = "ERRO: arquivo salvo não é um PDF válido"
                obs    = ("Provável causa: tipo de arquivo errado no diálogo 'Salvar como' "
                          "ou render incompleto antes de salvar (ver ESPERA_RENDER)")
                cont_erro += 1
                print(f"{prefixo} → ✘ FALHOU: PDF inválido "
                      f"[{time.time() - inicio_nota:.1f}s]")

        except Exception as e:
            status = f"ERRO: {e}"
            obs    = str(e)
            cont_erro += 1
            print(f"{prefixo} → ✘ ERRO: {e} [{time.time() - inicio_nota:.1f}s]")
            # em caso de erro, não confie no estado da tela: reabre na próxima
            selecao_pronta = False

        dados_relatorio.append({
            "linha": i, "ocorrencia": ocorrencia, "nota": num_nota,
            "tipo": "Serviço", "status": status, "obs": obs
        })

        if processadas % PARCIAL_A_CADA == 0:
            imprimir_parcial(processadas, total, cont_ok, cont_erro,
                             cont_material, inicio_geral)

    duracao_total = time.time() - inicio_geral
    print(f"\nProcessamento concluído em {formatar_duracao(duracao_total)} "
          f"({duracao_total / total:.1f}s por nota em média)." if total
          else "\nNenhuma nota para processar.")

    imprimir_relatorio_terminal(dados_relatorio)
    caminho_rel = gerar_relatorio_excel(PASTA_DESTINO, dados_relatorio)
    print(f"Planilha de relatório salva em: {caminho_rel}")


def gerar_relatorio_excel(pasta, dados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório NFS-e"

    fill_verde    = PatternFill("solid", fgColor="C6EFCE")
    fill_amarelo  = PatternFill("solid", fgColor="FFEB9C")
    fill_vermelho = PatternFill("solid", fgColor="FFC7CE")
    fill_titulo   = PatternFill("solid", fgColor="1F3864")
    fill_cabec    = PatternFill("solid", fgColor="2E75B6")

    fonte_titulo = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    fonte_cabec  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    fonte_normal = Font(name="Calibri", size=11)

    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    centro = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Relatório NFS-e  —  Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = fonte_titulo
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = centro
    ws.row_dimensions[1].height = 30

    cabecalhos = ["Linha", "Ocorrência", "Nota Fiscal", "Tipo", "Status PDF", "Observação"]
    for col, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=2, column=col, value=cab)
        cell.font = fonte_cabec
        cell.fill = fill_cabec
        cell.alignment = centro
        cell.border = borda
    ws.row_dimensions[2].height = 22

    for row_idx, d in enumerate(dados, 3):
        tipo   = d["tipo"]
        status = d["status"]
        if tipo == "Serviço" and status == "OK":
            fill = fill_verde
        elif tipo == "Serviço":
            fill = fill_amarelo
        else:
            fill = fill_vermelho

        valores = [d["linha"], d["ocorrencia"], d["nota"], tipo, status, d.get("obs", "")]
        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = fonte_normal
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="left" if col_idx == 6 else "center",
                vertical="center"
            )
            cell.border = borda

    for col, larg in enumerate([8, 18, 18, 12, 22, 35], 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    caminho = os.path.join(pasta, "_relatorio_nfse.xlsx")
    wb.save(caminho)
    return caminho


def imprimir_relatorio_terminal(dados):
    servicos  = [d for d in dados if d["tipo"] == "Serviço"]
    materiais = [d for d in dados if d["tipo"] == "Material"]
    ok        = [d for d in servicos if d["status"] == "OK"]
    erros     = [d for d in servicos if d["status"] != "OK"]

    print("\n" + "=" * 60)
    print("           RELATÓRIO FINAL — NFS-e")
    print("=" * 60)
    print(f"  Total de ocorrências           : {len(dados)}")
    print(f"  Serviço  (NFS-e encontrada)    : {len(servicos)}")
    print(f"    ✔  PDFs salvos com sucesso   : {len(ok)}")
    print(f"    ✘  Erros ao salvar PDF       : {len(erros)}")
    print(f"  Material (NFS-e não encontrada): {len(materiais)}")
    print("=" * 60)

    if materiais:
        print("\n  Ocorrências → MATERIAL (não encontradas no SAP):")
        for d in materiais:
            print(f"    Linha {d['linha']:>4} | Ocorrência {d['ocorrencia']} | Nota {d['nota']}")

    if erros:
        print("\n  Serviços com ERRO ao salvar PDF:")
        for d in erros:
            print(f"    Linha {d['linha']:>4} | Ocorrência {d['ocorrencia']} | {d['status']}")

    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()