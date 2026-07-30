import os
import time
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ========================= CONFIGURAÇÕES =========================
PASTA_DESTINO   = r"C:\Users\5829097\OneDrive\Onedrive - GPA\Área de Trabalho\NOTAS RECLASSIFICAÇÃO"
RELATORIO_EXCEL = os.path.join(PASTA_DESTINO, "_relatorio_nfse.xlsx")
PASTA_DIAGNOSTICO = os.path.join(PASTA_DESTINO, "_diagnostico_nix")

URL_NIX  = "https://nixweb.midassolutions.com.br/034/web/ItemPreOccurrence/"
BASE_URL = "https://nixweb.midassolutions.com.br"

PERFIL_NAVEGADOR = os.path.join(os.environ["USERPROFILE"],
                                "AppData", "Local", "nix_playwright_profile")

ESPERA_BUSCA  = 5.0
ESPERA_VIEWER = 4.0
TIMEOUT_RESPOSTA_PDF = 15000  # ms para aguardar a resposta do PDF após clicar
# =================================================================


def pdf_valido(caminho):
    """Verifica se o arquivo realmente começa com a assinatura de um PDF (%PDF-)."""
    try:
        if not os.path.exists(caminho) or os.path.getsize(caminho) == 0:
            return False
        with open(caminho, "rb") as f:
            assinatura = f.read(5)
        return assinatura == b"%PDF-"
    except Exception:
        return False


def salvar_diagnostico(nome_base, conteudo_bytes, content_type):
    """Quando o conteúdo baixado não é um PDF válido, salva uma cópia para inspeção
    manual (em vez de só descartar), junto com o content-type recebido."""
    os.makedirs(PASTA_DIAGNOSTICO, exist_ok=True)
    ext = ".html" if b"<html" in conteudo_bytes[:2000].lower() else ".bin"
    caminho_diag = os.path.join(PASTA_DIAGNOSTICO, f"{nome_base}{ext}")
    try:
        with open(caminho_diag, "wb") as f:
            f.write(conteudo_bytes)
        with open(caminho_diag + ".content-type.txt", "w", encoding="utf-8") as f:
            f.write(str(content_type))
    except Exception:
        pass
    return caminho_diag


def ler_materiais():
    if not os.path.exists(RELATORIO_EXCEL):
        print(f"ERRO: relatório não encontrado em:\n  {RELATORIO_EXCEL}")
        print("Rode o script principal (main.py) primeiro.")
        return []

    wb = openpyxl.load_workbook(RELATORIO_EXCEL, data_only=True)
    ws = wb.active

    materiais = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        linha, ocorrencia, nota, tipo, status, obs = row
        if str(tipo).strip().lower() == "material":
            materiais.append({
                "linha": linha,
                "ocorrencia": str(ocorrencia).strip(),
                "nota": str(nota).strip()
            })

    print(f"{len(materiais)} nota(s) de Material encontradas no relatório.")
    return materiais


def atualizar_relatorio(resultados):
    try:
        wb = openpyxl.load_workbook(RELATORIO_EXCEL)
    except Exception as e:
        print(f"\n⚠️  Feche o arquivo _relatorio_nfse.xlsx e tente novamente.\n{e}")
        return


    ws = wb.active
    fill_verde   = PatternFill("solid", fgColor="C6EFCE")
    fill_amarelo = PatternFill("solid", fgColor="FFEB9C")
    fonte_normal = Font(name="Calibri", size=11)
    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    for row_idx in range(3, ws.max_row + 1):
        linha_planilha = ws.cell(row=row_idx, column=1).value
        if linha_planilha in resultados:
            r = resultados[linha_planilha]
            fill = fill_verde if r["status"] == "OK" else fill_amarelo
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = fill
                ws.cell(row=row_idx, column=col).font = fonte_normal
                ws.cell(row=row_idx, column=col).border = borda
            ws.cell(row=row_idx, column=5).value = r["status"]
            ws.cell(row=row_idx, column=6).value = r["obs"]

    try:
        wb.save(RELATORIO_EXCEL)
        print("Relatório atualizado com sucesso.")
    except PermissionError:
        print("\n⚠️  Feche o _relatorio_nfse.xlsx e tente novamente.")


def encontrar_botao_download(pagina, timeout_ms):
    """
    O visualizador de PDF pode aparecer de duas formas:
    1) Navegação cheia da aba/página para o PDF (o botão fica no documento principal).
    2) Um modal/preview embutido via <iframe> na própria página (comum quando o
       clique não abre nova aba nem navega a página principal) — nesse caso o
       botão #download fica DENTRO do iframe, e page.locator() sozinho não
       consegue enxergar isso porque não atravessa fronteiras de iframe.
    Aqui tentamos achar o botão primeiro na página principal e, se não achar,
    procuramos em cada iframe carregado até o timeout.
    """
    try:
        btn = pagina.locator("#download")
        btn.wait_for(state="visible", timeout=2000)
        return btn
    except PlaywrightTimeout:
        pass

    deadline = time.time() + (timeout_ms / 1000)
    while time.time() < deadline:
        for frame in pagina.frames:
            try:
                btn = frame.locator("#download")
                if btn.count() > 0:
                    btn.wait_for(state="visible", timeout=1000)
                    return btn
            except Exception:
                continue
        pagina.wait_for_timeout(300)

    # não achou - loga as URLs dos frames pra facilitar diagnóstico
    try:
        urls_frames = [f.url for f in pagina.frames]
        print(f"    (debug) frames encontrados na página: {urls_frames}")
    except Exception:
        pass

    return None


def salvar_diagnostico_visual(pagina, nome_base):
    """Salva um screenshot da página no momento da falha, pra facilitar
    identificar visualmente o que apareceu (modal, iframe, erro, etc.)."""
    os.makedirs(PASTA_DIAGNOSTICO, exist_ok=True)
    caminho_print = os.path.join(PASTA_DIAGNOSTICO, f"{nome_base}_falha.png")
    try:
        pagina.screenshot(path=caminho_print, timeout=5000)
    except Exception:
        caminho_print = None
    return caminho_print


def baixar_pdf_via_botao_viewer(page, context, botao_visualizar, caminho_destino,
                                 timeout_ms=TIMEOUT_RESPOSTA_PDF):
    """
    Clica no botão Visualizar, que abre o visualizador de PDF do navegador
    (PDF.js) — em nova aba, navegação na mesma aba, ou num iframe/modal
    embutido na própria página — e então clica no botão de Download desse
    visualizador (id="download"), capturando o download real via Playwright.

    Retorna True se salvou com sucesso, False caso contrário.
    """
    pagina_viewer = page

    # o clique pode abrir uma nova aba com o visualizador
    try:
        with context.expect_page(timeout=5000) as new_page_info:
            botao_visualizar.click()
        pagina_viewer = new_page_info.value
        pagina_viewer.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
    except PlaywrightTimeout:
        # não abriu nova aba - o clique já foi feito; o PDF costuma abrir num
        # modal (Bootstrap) embutido na própria página, com um iframe dentro.
        # Esperamos o modal aparecer antes de procurar o botão de download,
        # pra não competir com a animação de abertura.
        try:
            page.locator("#divModalTitle").wait_for(state="visible", timeout=5000)
        except Exception:
            pass
        pagina_viewer = page

    try:
        btn_download = encontrar_botao_download(pagina_viewer, timeout_ms)
        if btn_download is None:
            caminho_print = salvar_diagnostico_visual(
                pagina_viewer, os.path.splitext(os.path.basename(caminho_destino))[0]
            )
            raise Exception(
                "Botão de download do visualizador não encontrado (nem na página "
                "nem em iframes)." + (f" Screenshot salvo em: {caminho_print}" if caminho_print else "")
            )

        with pagina_viewer.expect_download(timeout=timeout_ms) as download_info:
            btn_download.click()
        download = download_info.value
        download.save_as(caminho_destino)
        return True
    finally:
        if pagina_viewer is not page:
            try:
                pagina_viewer.close()
            except Exception:
                pass


def processar_materiais_nix(materiais):
    os.makedirs(PERFIL_NAVEGADOR, exist_ok=True)
    resultados = {}

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            PERFIL_NAVEGADOR,
            headless=False,
            accept_downloads=True,
            channel="msedge",
            ignore_https_errors=True,
            args=["--start-maximized"],
            no_viewport=True
        )

        page = context.new_page() if not context.pages else context.pages[0]

        page.goto(URL_NIX, wait_until="domcontentloaded")
        time.sleep(2)

        # trata login automaticamente
        if "Login" in page.title() or "Account" in page.url:
            print("Tela de login detectada — clicando em Entrar...")
            try:
                btn_entrar = page.locator("#btn-modal")
                btn_entrar.wait_for(state="visible", timeout=5000)
                btn_entrar.click()
                time.sleep(2)
                page.wait_for_url(
                    lambda url: "Login" not in url and "Account" not in url,
                    timeout=15000
                )
                print("Login realizado com sucesso!")
            except Exception as e:
                print(f"Não consegui clicar em Entrar automaticamente: {e}")
                print("Faça o login manualmente e pressione ENTER...")
                input()
            page.goto(URL_NIX, wait_until="domcontentloaded")
            time.sleep(2)

        for item in materiais:
            ocorrencia = item["ocorrencia"]
            num_nota   = item["nota"]
            linha      = item["linha"]
            nome_pdf   = f"{num_nota}.pdf"
            caminho    = os.path.join(PASTA_DESTINO, nome_pdf)

            print(f"\nLinha {linha}: ocorrência {ocorrencia} → buscando no NIX...")

            try:
                if URL_NIX not in page.url:
                    page.goto(URL_NIX, wait_until="domcontentloaded")
                    time.sleep(1.5)

                # preenche ocorrência e limpa datas (datas às vezes vêm desabilitadas)
                campo_ocor = page.locator("#GaiaOccurenceId")
                campo_ocor.wait_for(state="visible", timeout=10000)
                campo_ocor.fill(ocorrencia)
                for seletor in ("#InitialDateCreate", "#EndDateCreate"):
                    campo_data = page.locator(seletor)
                    try:
                        if campo_data.is_enabled(timeout=2000):
                            campo_data.fill("")
                    except Exception:
                        pass
                time.sleep(0.3)

                campo_ocor.press("Enter")
                time.sleep(ESPERA_BUSCA)

                # verifica se encontrou resultado
                btn_visualizar = page.locator("a.MvcGridSearchButton").first
                if not btn_visualizar.is_visible():
                    print(f"  → Ocorrência {ocorrencia} não encontrada no NIX.")
                    resultados[linha] = {
                        "status": "ERRO: não encontrada no NIX",
                        "obs": "Ocorrência não retornou resultado no NIX"
                    }
                    continue

                # clica em Visualizar, depois no botão de Download do viewer,
                # e captura o download real do navegador
                try:
                    baixar_pdf_via_botao_viewer(page, context, btn_visualizar, caminho)
                except (PlaywrightTimeout, Exception) as e:
                    print(f"  → [FALHOU] {e}")
                    resultados[linha] = {
                        "status": "ERRO: viewer/download",
                        "obs": str(e)
                    }
                    continue

                if pdf_valido(caminho):
                    print(f"  → {nome_pdf} salvo com sucesso [OK]")
                    resultados[linha] = {"status": "OK", "obs": caminho}
                else:
                    if os.path.exists(caminho):
                        os.remove(caminho)
                    print(f"  → [FALHOU] arquivo baixado não é um PDF válido")
                    resultados[linha] = {
                        "status": "ERRO: arquivo baixado não é PDF válido",
                        "obs": "O download ocorreu mas o conteúdo salvo não tem assinatura de PDF"
                    }

            except Exception as e:
                print(f"  → [ERRO] {e}")
                resultados[linha] = {"status": f"ERRO: {e}", "obs": str(e)}

            page.goto(URL_NIX, wait_until="domcontentloaded")
            time.sleep(1.5)

        context.close()

    return resultados


def imprimir_resumo(resultados):
    ok    = [r for r in resultados.values() if r["status"] == "OK"]
    erros = [r for r in resultados.values() if r["status"] != "OK"]

    print("\n" + "=" * 60)
    print("        RELATÓRIO FINAL — NIX Material")
    print("=" * 60)
    print(f"  Total processado      : {len(resultados)}")
    print(f"  ✔  Salvos com sucesso : {len(ok)}")
    print(f"  ✘  Erros              : {len(erros)}")
    print("=" * 60 + "\n")


def main():
    os.makedirs(PASTA_DESTINO, exist_ok=True)

    materiais = ler_materiais()
    if not materiais:
        return

    resultados = processar_materiais_nix(materiais)
    atualizar_relatorio(resultados)
    imprimir_resumo(resultados)
    print(f"Relatório em: {RELATORIO_EXCEL}")


if __name__ == "__main__":
    main()